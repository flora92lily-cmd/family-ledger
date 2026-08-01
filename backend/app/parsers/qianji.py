"""钱迹（QianJi）账本解析器（支持新版 xlsx 与旧版 CSV）

钱迹导出格式：
- 新版：xlsx，表头新增“账本”列，时间格式为 YYYY-MM-DD HH:MM:SS
- 旧版：UTF-8 BOM CSV，时间格式为 M/D/YYYY H:MM
- 业务表头：ID, 时间, [账本], 分类, 二级分类, 类型, 金额, 币种, 账户1, 账户2, 备注,
        已报销, 手续费, 优惠券, 记账者, 账单标记, 标签, 账单图片, 关联账单
- 类型取值：支出 / 收入 / 还款 / 转账 / 报销 / 报销记录
  - 支出/收入 → 普通交易
  - 报销 → 可报销支出（is_reimbursable=True），账户1=扣款账户
  - 报销记录 → 报销到账记录（不建 Transaction，建 ReimbursementRecord），
              账户1=到账账户，关联账单=原报销支出的 ID
  - 转账/还款 等双账户场景 → 由账户1+账户2 是否同时非空判定（见下），不依赖类型字符串
- 转账识别：任何一行 账户1 与 账户2 同时非空 → type=transfer，
            账户1=转出，账户2=转入（覆盖 类型=转账/还款/借出/借入 等）
- 已报销列：值为"是" 表示该可报销支出已被报销；空表示待报销
- 标签：作为家庭成员/项目等多维度标识

分类匹配策略：
- description = 备注列（备注空则空白；不再用分类名兜底，避免列表显示分类名）
- source_category_name = 钱迹二级分类（categorizer 优先按名直接匹配）
- source_parent_category_name = 钱迹一级分类（次选 fallback）
"""
import csv
import io
from datetime import datetime
from app.parsers.base import BaseParser, ParsedTransaction, ParsedReimbursement


class QianjiParser(BaseParser):
    name = "qianji"

    def parse(self, file_bytes: bytes, filename: str = "") -> list:
        """返回混合列表：ParsedTransaction + ParsedReimbursement"""
        if filename.lower().endswith(".xlsx") or file_bytes[:2] == b"PK":
            return self._parse_xlsx(file_bytes)
        return self._parse_csv(file_bytes)

    def _parse_xlsx(self, file_bytes: bytes) -> list:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                return []

            headers = [self._cell_text(cell) for cell in header_row]
            return self._parse_rows(rows, headers)
        finally:
            workbook.close()

    def _parse_csv(self, file_bytes: bytes) -> list:
        text = self._decode(file_bytes)
        lines = text.splitlines()
        if not lines:
            return []

        reader = csv.reader(lines)
        headers = [h.strip() for h in next(reader)]
        return self._parse_rows(reader, headers)

    def _parse_rows(self, rows, headers: list[str]) -> list:
        col = {
            "id": self._find_col(headers, ["ID"]),
            "time": self._find_col(headers, ["时间"]),
            "category": self._find_col(headers, ["分类"]),
            "subcategory": self._find_col(headers, ["二级分类"]),
            "type": self._find_col(headers, ["类型"]),
            "amount": self._find_col(headers, ["金额"]),
            "account": self._find_col(headers, ["账户1"]),
            "account2": self._find_col(headers, ["账户2"]),
            "note": self._find_col(headers, ["备注"]),
            "reimbursed": self._find_col(headers, ["已报销"]),
            "tag": self._find_col(headers, ["标签"]),
            "linked": self._find_col(headers, ["关联账单"]),
        }

        items: list = []
        for row in rows:
            if not row or len(row) < 6:
                continue
            try:
                item = self._parse_row(row, col)
                if item:
                    items.append(item)
            except (ValueError, IndexError):
                continue
        return items

    def _cell_text(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _decode(self, b: bytes) -> str:
        for enc in ("utf-8-sig", "utf-8", "gbk"):
            try:
                return b.decode(enc)
            except UnicodeDecodeError:
                continue
        return b.decode("utf-8", errors="ignore")

    def _find_col(self, headers: list[str], candidates: list[str]) -> int:
        # 严格相等优先，否则模糊
        for cand in candidates:
            for i, h in enumerate(headers):
                if h == cand:
                    return i
        for cand in candidates:
            for i, h in enumerate(headers):
                if cand in h:
                    return i
        return -1

    def _parse_row(self, row: list | tuple, col: dict):
        def safe(key):
            idx = col.get(key, -1)
            return self._cell_text(row[idx]) if 0 <= idx < len(row) else ""

        type_str = safe("type")

        # 解析通用字段
        amount_str = safe("amount").replace(",", "").strip()
        try:
            amount = abs(float(amount_str))
        except ValueError:
            return None
        if amount == 0:
            return None

        time_str = safe("time")
        date_part = time_str.split()[0] if " " in time_str else time_str
        txn_date = None
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                txn_date = datetime.strptime(date_part, fmt).date()
                break
            except ValueError:
                continue
        if not txn_date:
            return None

        row_id = safe("id")
        linked_id = safe("linked")
        payment_method = safe("account")
        to_payment_method = safe("account2")
        note_text = safe("note")  # 钱迹"备注"列

        # 分支 1：报销记录 → ParsedReimbursement
        if type_str == "报销记录":
            return ParsedReimbursement(
                amount=amount,
                date=txn_date,
                payment_method=payment_method,
                note=note_text,
                external_id=row_id,
                linked_external_id=linked_id,
                raw=",".join(self._cell_text(cell) for cell in row)[:200],
            )

        # 分支 2：账户1 + 账户2 同时非空 → transfer（覆盖 类型=转账/还款/借出/借入 等）
        # 不依赖"类型"字符串，因为钱迹的"还款"等也是双账户语义
        if payment_method and to_payment_method:
            sub_cat = safe("subcategory")
            parent_cat = safe("category")
            tag_str = safe("tag")
            tags = [t.strip() for t in tag_str.split(",") if t.strip()] if tag_str else []
            # description 兜底：备注 > 二级分类 > 类型字符串(还款/转账) > 一级分类
            # 一级分类常为"其它"无信息量，放最后；类型字符串("还款"/"转账")更直观
            meaningful = lambda s: s and s not in ("其它", "其他")
            desc = (
                note_text
                or (sub_cat if meaningful(sub_cat) else "")
                or (type_str if type_str else "")
                or (parent_cat if parent_cat else "")
                or "转账"
            )
            return ParsedTransaction(
                amount=amount,
                type="transfer",
                date=txn_date,
                description=desc,
                counterparty="",
                payment_method=payment_method,
                to_payment_method=to_payment_method,
                tags=tags,
                # 转账不归类（category_id 由用户在 APP 内决定，通常留空）
                source_category_name=None,
                source_parent_category_name=None,
                external_id=row_id,
                raw=",".join(self._cell_text(cell) for cell in row)[:200],
            )

        # 分支 3：普通交易 / 报销支出
        if type_str == "支出":
            txn_type = "expense"
            is_reimbursable = False
        elif type_str == "收入":
            txn_type = "income"
            is_reimbursable = False
        elif type_str == "报销":
            txn_type = "expense"
            is_reimbursable = True
        else:
            # 类型=转账/还款 但缺一边账户，无法成立 transfer → 跳过
            return None

        # description = 备注（首页列表显示这一字段）
        # 分类名通过 source_category_name / source_parent_category_name 传给 categorizer
        sub_cat = safe("subcategory")
        parent_cat = safe("category")
        description = note_text  # 备注为空就显示空白

        # 标签
        tag_str = safe("tag")
        tags = [t.strip() for t in tag_str.split(",") if t.strip()] if tag_str else []

        # 报销相关字段
        reimbursable_amount = 0.0
        reimbursement_status = "none"
        if is_reimbursable:
            reimbursable_amount = amount  # 乐观估计=实付金额（后续如果有关联报销记录，save 阶段会改写）
            reimbursed_flag = safe("reimbursed")
            reimbursement_status = "done" if reimbursed_flag == "是" else "pending"

        return ParsedTransaction(
            amount=amount,
            type=txn_type,
            date=txn_date,
            description=description,
            counterparty="",
            payment_method=payment_method,
            tags=tags,
            source_category_name=sub_cat or None,
            source_parent_category_name=parent_cat or None,
            is_reimbursable=is_reimbursable,
            reimbursable_amount=reimbursable_amount,
            reimbursement_status=reimbursement_status,
            external_id=row_id,
            raw=",".join(self._cell_text(cell) for cell in row)[:200],
        )
