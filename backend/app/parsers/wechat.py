"""微信支付账单解析器（支持 xlsx 和 csv）

微信支付的官方导出格式是 xlsx：
- 前 16 行是说明文字
- 第 17 行（index 16）是表头：交易时间, 交易类型, 交易对方, 商品, 收/支, 金额(元),
  支付方式, 当前状态, 交易单号, 商户单号, 备注

收/支 取值：支出 / 收入 / 中性交易（"/" 等）
中性交易：充值、提现、理财通、信用卡还款 → 跳过
特例：零钱通存入/转出 → 转账（type=transfer），保留导入

description 拼接规则：
- 商品 ∈ {"/", "", "-"} → description = 交易对方
- 交易对方为空 → description = 商品
- 都有 → description = "交易对方，商品"
"""
import csv
import io
from datetime import datetime
from app.parsers.base import BaseParser, ParsedTransaction


class WechatParser(BaseParser):
    name = "wechat"

    def parse(self, file_bytes: bytes, filename: str = "") -> list[ParsedTransaction]:
        # 优先按 xlsx 解析
        if filename.lower().endswith((".xlsx", ".xls")) or file_bytes[:2] == b"PK":
            return self._parse_xlsx(file_bytes)
        return self._parse_csv(file_bytes)

    # ========== xlsx ==========
    def _parse_xlsx(self, file_bytes: bytes) -> list[ParsedTransaction]:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        header_idx = self._find_header_row(rows)
        if header_idx < 0:
            return []

        headers = [str(c or "").strip() for c in rows[header_idx]]
        col = {
            "time": self._find_col(headers, ["交易时间"]),
            "category": self._find_col(headers, ["交易类型"]),
            "counterparty": self._find_col(headers, ["交易对方"]),
            "description": self._find_col(headers, ["商品"]),
            "type": self._find_col(headers, ["收/支"]),
            "amount": self._find_col(headers, ["金额(元)", "金额"]),
            "status": self._find_col(headers, ["当前状态"]),
            "payment_method": self._find_col(headers, ["支付方式"]),
        }

        transactions = []
        for row in rows[header_idx + 1:]:
            if not row or all(c is None or not str(c).strip() for c in row):
                continue
            try:
                txn = self._parse_xlsx_row(row, col)
                if txn:
                    transactions.append(txn)
            except (ValueError, IndexError):
                continue
        return transactions

    def _find_header_row(self, rows: list) -> int:
        for i, row in enumerate(rows):
            cells = [str(c or "") for c in row]
            joined = " ".join(cells)
            if "交易时间" in joined and "收/支" in joined:
                return i
        return -1

    def _parse_xlsx_row(self, row: tuple, col: dict) -> ParsedTransaction | None:
        def safe(key):
            idx = col.get(key, -1)
            if 0 <= idx < len(row):
                v = row[idx]
                return str(v).strip() if v is not None else ""
            return ""

        type_str = safe("type")
        category = safe("category")  # 交易类型列（如"零钱通存入"/"零钱通转出"/"商户消费"）
        product = safe("description")  # 商品列
        counterparty = safe("counterparty")

        # 9.6 零钱通转账识别：收/支="/" 或中性 + 交易类型/商品含"零钱通"
        is_zqt = "零钱通" in category or "零钱通" in product
        if "收入" in type_str:
            txn_type = "income"
        elif "支出" in type_str:
            txn_type = "expense"
        elif is_zqt:
            txn_type = "transfer"
        else:
            # 其他中性交易（充值/提现/理财通/信用卡还款等）跳过
            return None

        amount_str = safe("amount").replace("¥", "").replace(",", "").strip()
        try:
            amount = abs(float(amount_str))
        except ValueError:
            return None
        if amount == 0:
            return None

        time_str = safe("time")
        try:
            # 可能是 datetime 对象（openpyxl 自动解析）也可能是字符串
            if " " in time_str:
                txn_date = datetime.strptime(time_str.split()[0], "%Y-%m-%d").date()
            else:
                txn_date = datetime.strptime(time_str, "%Y-%m-%d").date()
        except (ValueError, IndexError):
            return None

        # 9.4 description 拼接：交易对方 + 商品（用逗号），转账类型直接用交易类型
        if txn_type == "transfer":
            desc = category or product or "转账"
        else:
            product_clean = product if product not in ("/", "", "-") else ""
            if counterparty and product_clean:
                desc = f"{counterparty}，{product_clean}"
            elif counterparty:
                desc = counterparty
            elif product_clean:
                desc = product_clean
            else:
                desc = category  # 兜底用交易类型

        return ParsedTransaction(
            amount=amount,
            type=txn_type,
            date=txn_date,
            description=desc,
            counterparty=counterparty,
            payment_method=safe("payment_method"),
            raw=" | ".join(str(c or "") for c in row)[:200],
        )

    # ========== csv ==========
    def _parse_csv(self, file_bytes: bytes) -> list[ParsedTransaction]:
        text = self._decode(file_bytes)
        lines = text.splitlines()
        header_idx = self._find_csv_header(lines)
        if header_idx < 0:
            return []

        reader = csv.reader(lines[header_idx:])
        headers = [h.strip() for h in next(reader)]
        col = {
            "time": self._find_col(headers, ["交易时间"]),
            "category": self._find_col(headers, ["交易类型"]),
            "counterparty": self._find_col(headers, ["交易对方"]),
            "description": self._find_col(headers, ["商品"]),
            "type": self._find_col(headers, ["收/支"]),
            "amount": self._find_col(headers, ["金额(元)", "金额"]),
            "status": self._find_col(headers, ["当前状态"]),
            "payment_method": self._find_col(headers, ["支付方式"]),
        }

        transactions = []
        for row in reader:
            if not row or len(row) < 5:
                continue
            try:
                txn = self._parse_xlsx_row(row, col)  # 兼容 list 行
                if txn:
                    transactions.append(txn)
            except (ValueError, IndexError):
                continue
        return transactions

    def _decode(self, b: bytes) -> str:
        for enc in ("utf-8-sig", "utf-8", "gbk"):
            try:
                return b.decode(enc)
            except UnicodeDecodeError:
                continue
        return b.decode("utf-8", errors="ignore")

    def _find_csv_header(self, lines: list[str]) -> int:
        for i, line in enumerate(lines):
            if "交易时间" in line and "收/支" in line:
                return i
        return -1

    def _find_col(self, headers: list[str], candidates: list[str]) -> int:
        for cand in candidates:
            for i, h in enumerate(headers):
                if cand in h:
                    return i
        return -1
